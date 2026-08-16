"""
Скрипт экспорта полученных дропов в Excel для участников с class_id=2.
Логика: для каждого дропа смотрим, за какую задачу он выдан (task_id), находим
решение этой задачи в boss_task_solution; если у решения class_id == 2,
записываем в Excel имя получателя из boss_task_solution.user_name.
Колонки: id участника, имя участника, название дропа.
Игнорирует записи с дропами, содержащими "пустой", "ничегошеньки", "сундук" (без учёта регистра).

Установка: pip install openpyxl
"""
from __future__ import annotations

import contextlib
import io
import os
import sys
from datetime import datetime

# Слова для исключения (подстрока в названии дропа, без учёта регистра)
SKIP_DROP_PATTERNS = ("пустой", "пусто", "ничего", "ничегошеньки", "сундук")


def _get_solution_for_reward(reward, BossTaskSolution):
    """
    Возвращает решение задачи, за которое выдан дроп.
    Если у награды указан task_id — ищем решение по (boss_id, task_id, user_id).
    Если task_id нет — берём последнее правильное решение по этому боссу до момента дропа.
    Возвращает BossTaskSolution или None.
    """
    if reward.task_id is not None:
        return (
            BossTaskSolution.query.filter_by(
                boss_id=reward.boss_id,
                task_id=reward.task_id,
                user_id=reward.user_id,
                is_correct=True,
            )
            .order_by(BossTaskSolution.solved_at.desc())
            .first()
        )
    return (
        BossTaskSolution.query.filter_by(
            boss_id=reward.boss_id, user_id=reward.user_id, is_correct=True
        )
        .filter(BossTaskSolution.solved_at <= reward.received_at)
        .order_by(BossTaskSolution.solved_at.desc())
        .first()
    )


def _should_skip_drop(drop_name: str | None) -> bool:
    """True, если дроп нужно пропустить (содержит запрещённые слова)."""
    if not drop_name:
        return False
    name_lower = drop_name.lower()
    return any(p in name_lower for p in SKIP_DROP_PATTERNS)


def _safe_output_path(path: str) -> str:
    """
    Если файл уже открыт (например, в Excel) и заблокирован, не падаем,
    а создаём новый файл рядом с таймстампом.
    """
    try:
        with open(path, "a", encoding="utf-8"):
            return path
    except PermissionError:
        root, ext = os.path.splitext(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{root}_{ts}{ext}"


def main() -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        print("Ошибка: требуется openpyxl. Установите: pip install openpyxl")
        sys.exit(1)

    os.makedirs("exports", exist_ok=True)
    out_path = _safe_output_path("exports/drops_class2.xlsx")

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        from app import app, db, BossDropReward, BossTaskSolution  # type: ignore[import-untyped]

    with app.app_context():
        rewards = (
            BossDropReward.query.order_by(
                BossDropReward.user_id.asc(), BossDropReward.received_at.asc()
            )
            .all()
        )

        rows = []
        for r in rewards:
            solution = _get_solution_for_reward(r, BossTaskSolution)
            if not solution or solution.class_id != 2:
                continue
            drop_name = r.drop.name if r.drop else ""
            if _should_skip_drop(drop_name):
                continue
            # Имя получателя — как в таблице boss_task_solution
            user_name = solution.user_name or ""
            rows.append((r.user_id, user_name, drop_name))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Дропы class_id=2"

        headers = ["id участника", "имя участника", "название дропа"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for row_idx, (uid, uname, dname) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=uid)
            ws.cell(row=row_idx, column=2, value=uname)
            ws.cell(row=row_idx, column=3, value=dname)

        wb.save(out_path)

    print(f"OK: экспортировано {len(rows)} записей в {out_path}")


if __name__ == "__main__":
    main()
